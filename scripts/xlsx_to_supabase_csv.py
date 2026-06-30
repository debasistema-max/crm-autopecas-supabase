import csv
import re
import sys
import unicodedata
from pathlib import Path
from openpyxl import load_workbook


def clean(value):
    return "" if value is None else str(value).strip()


def product_code(value):
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    text = clean(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def norm(value):
    text = clean(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = (
        text.replace("ç", "c")
        .replace("ã", "a")
        .replace("á", "a")
        .replace("à", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("õ", "o")
        .replace("ô", "o")
        .replace("ú", "u")
    )
    return re.sub(r"[^a-z0-9]+", "", text)


def number(value):
    text = clean(value)
    if not text:
        return "0"
    if isinstance(value, (int, float)):
        return str(value)
    text = re.sub(r"[^\d,.-]", "", text)
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return str(float(text))
    except ValueError:
        return "0"


def rows(ws):
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = [norm(v) for v in next(iterator)]
    except StopIteration:
        return []
    output = []
    for row in iterator:
        obj = {}
        for idx, header in enumerate(headers):
            if header:
                obj[header] = row[idx] if idx < len(row) else ""
        if any(clean(v) for v in obj.values()):
            output.append(obj)
    return output


def first_sheet(wb, names):
    wanted = {norm(name) for name in names}
    for ws in wb.worksheets:
        if norm(ws.title) in wanted:
            return ws
    return None


def pick(obj, aliases):
    for alias in aliases:
        key = norm(alias)
        if key in obj:
            return obj[key]
    return ""


def write_csv(path, headers, data):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for item in data:
            writer.writerow({h: item.get(h, "") for h in headers})


def build_products(wb):
    base_ws = first_sheet(wb, ["BASE_UNIFICADA", "MATRIZ", "IMPORTACAO DESCRICAO", "IMPORTAÇÃO DESCRIÇÃO"])
    sp_ws = first_sheet(wb, ["PRECO_SP", "TABELA SP", "IMPORTACAO SP", "IMPORTAÇÃO SP"])
    pr_ws = first_sheet(wb, ["PRECO_PR", "TABELA PR", "IMPORTACAO PR", "IMPORTAÇÃO PR"])
    catalog_ws = first_sheet(wb, ["CATALOGO_PESQUISA"])

    products = {}
    if base_ws:
        for row in rows(base_ws):
            code = product_code(pick(row, ["codigo", "codigo ips", "cod", "cód."]))
            if not code:
                continue
            products[code] = {
                "codigo": code,
                "descricao": clean(pick(row, ["descricao", "descrição"])),
                "marca": clean(pick(row, ["marca"])),
                "aplicacao": clean(pick(row, ["aplicacao", "aplicação"])),
                "ano": clean(pick(row, ["ano"])),
                "ipi": number(pick(row, ["ipi"])),
                "preco_sem_imposto": number(pick(row, ["preco_sem_imposto", "pr.unit.", "preco sem imposto"])),
                "estoque": clean(pick(row, ["estoque", "disp. geral", "qtde"])),
                "estoque_quantidade": number(pick(row, ["estoque", "disp. geral", "qtde"])),
                "preco_sp": number(pick(row, ["preco_sp", "preco sp", "pr.unit.(c/i)", "pr.apos desc"])),
                "preco_pr": number(pick(row, ["preco_pr", "preco pr", "pr.unit.(c/i)", "pr.apos desc"])),
                "status_estoque": clean(pick(row, ["status_estoque"])),
                "status_cadastro": clean(pick(row, ["status_cadastro"])),
                "url_imagem": clean(pick(row, ["url_imagem", "imagem"])),
            }

    for ws, field in [(sp_ws, "preco_sp"), (pr_ws, "preco_pr")]:
        if not ws:
            continue
        for row in rows(ws):
            code = product_code(pick(row, ["codigo", "cod", "cód.", "codigo ips"]))
            if not code:
                continue
            products.setdefault(code, {"codigo": code})
            products[code][field] = number(pick(row, [field, "pr.unit.(c/i)", "pr.apos desc", "preco", "valor"]))
            if not products[code].get("descricao"):
                products[code]["descricao"] = clean(pick(row, ["descricao", "descrição"]))
            if not products[code].get("estoque"):
                products[code]["estoque"] = clean(pick(row, ["disp. geral", "estoque", "qtde"]))
                products[code]["estoque_quantidade"] = number(products[code]["estoque"])

    if catalog_ws:
        for row in rows(catalog_ws):
            code = product_code(pick(row, ["codigo", "cod", "sku"]))
            if not code:
                continue
            products.setdefault(code, {"codigo": code})
            for field in ["grupo", "categoria", "montadora", "detalhes", "marca", "ano", "oem", "similar"]:
                value = clean(pick(row, [field]))
                if value:
                    products[code][field] = value
            value = clean(pick(row, ["descricao", "descrição"]))
            if value and not products[code].get("descricao"):
                products[code]["descricao"] = value
            value = clean(pick(row, ["aplicacao", "aplicação"]))
            if value and not products[code].get("aplicacao"):
                products[code]["aplicacao"] = value

    headers = [
        "codigo", "descricao", "marca", "aplicacao", "ano", "ipi", "preco_sem_imposto",
        "estoque", "estoque_quantidade", "preco_sp", "preco_pr", "status_estoque",
        "status_cadastro", "url_imagem", "grupo", "categoria", "montadora", "detalhes",
        "oem", "similar"
    ]
    return headers, list(products.values())


def simple_table(wb, sheet_names, headers, aliases):
    ws = first_sheet(wb, sheet_names)
    if not ws:
        return []
    data = []
    for row in rows(ws):
        item = {}
        for header in headers:
            item[header] = clean(pick(row, aliases.get(header, [header])))
        if any(item.values()):
            data.append(item)
    return data


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: xlsx_to_supabase_csv.py input.xlsx output_dir")
    input_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(input_path, read_only=True, data_only=True)

    product_headers, products = build_products(wb)
    write_csv(output_dir / "products.csv", product_headers, products)

    carriers_headers = ["legacy_id", "cnpj", "nome", "telefone", "endereco", "ativo"]
    carriers = simple_table(wb, ["TRANSPORTADORAS"], carriers_headers, {
        "legacy_id": ["id_transportadora"],
        "nome": ["nome", "transportadora"],
    })
    write_csv(output_dir / "carriers.csv", carriers_headers, carriers)

    terms_headers = ["legacy_id", "descricao", "ativo"]
    terms = simple_table(wb, ["PRAZOS"], terms_headers, {"legacy_id": ["id_prazo"]})
    write_csv(output_dir / "payment_terms.csv", terms_headers, terms)

    print(f"products: {len(products)}")
    print(f"carriers: {len(carriers)}")
    print(f"payment_terms: {len(terms)}")


if __name__ == "__main__":
    main()
